"""
Sebastian Pre-processor — อ่านไฟล์ TOR (PDF/Excel) แล้วแปลงเป็น text กระชับ
ก่อนส่งให้ AI วิเคราะห์ (ไม่ใช้ AI ในขั้นนี้)

Input:  ไฟล์ PDF/Excel ใน downloads/
Output: dict ที่มี cleaned_text พร้อมส่ง AI
"""

import re
import json
from pathlib import Path
from typing import Optional

# ---- DEPS ----
# pip install pymupdf openpyxl

MAX_CHARS = 8000  # ตัด text ให้ไม่เกินนี้ก่อนส่ง AI

SECTION_KEYWORDS = [
    "ขอบเขต", "ปริมาณงาน", "ข้อกำหนด", "สเปก", "specification",
    "วัสดุ", "คอนกรีต", "เหล็ก", "ถนน", "สะพาน", "ท่อ",
    "ระยะทาง", "กว้าง", "ยาว", "หนา", "มิติ",
    "งบประมาณ", "วงเงิน", "ราคากลาง",
    "ระยะเวลา", "กำหนดแล้วเสร็จ",
]


# ================================================================
# PDF READER
# ================================================================

def read_pdf(filepath: Path) -> str:
    try:
        import fitz  # pymupdf
    except ImportError:
        return f"[ERROR: ต้องติดตั้ง pymupdf — pip install pymupdf] path={filepath}"

    text_parts = []
    try:
        doc = fitz.open(str(filepath))
        for page_num, page in enumerate(doc):
            text = page.get_text("text")
            if text.strip():
                text_parts.append(f"[หน้า {page_num+1}]\n{text}")
            if len("\n".join(text_parts)) > MAX_CHARS * 3:
                break  # อ่านพอแล้ว ไม่ต้องอ่านทั้งหมด
        doc.close()
    except Exception as e:
        return f"[ERROR reading PDF: {e}]"

    return "\n".join(text_parts)


# ================================================================
# EXCEL READER
# ================================================================

def read_excel(filepath: Path) -> str:
    try:
        import openpyxl
    except ImportError:
        return f"[ERROR: ต้องติดตั้ง openpyxl — pip install openpyxl] path={filepath}"

    text_parts = []
    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
        for sheet_name in wb.sheetnames[:3]:  # max 3 sheets
            ws = wb[sheet_name]
            text_parts.append(f"[Sheet: {sheet_name}]")
            for row in ws.iter_rows(max_row=200, values_only=True):
                row_text = " | ".join(str(v) for v in row if v is not None)
                if row_text.strip():
                    text_parts.append(row_text)
            if len("\n".join(text_parts)) > MAX_CHARS * 2:
                break
        wb.close()
    except Exception as e:
        return f"[ERROR reading Excel: {e}]"

    return "\n".join(text_parts)


# ================================================================
# TEXT CLEANER
# ================================================================

def clean_text(raw: str) -> str:
    # ลบ whitespace ซ้ำ
    text = re.sub(r'\n{3,}', '\n\n', raw)
    text = re.sub(r'[ \t]{2,}', ' ', text)
    text = re.sub(r'\r\n', '\n', text)

    # ลบ headers/footers ซ้ำ (บรรทัดสั้นๆ ที่ซ้ำกัน)
    lines = text.split('\n')
    seen_short = {}
    filtered = []
    for line in lines:
        stripped = line.strip()
        if len(stripped) < 15:
            seen_short[stripped] = seen_short.get(stripped, 0) + 1
            if seen_short[stripped] > 3:
                continue  # skip repeated short lines (headers/footers)
        filtered.append(line)

    return '\n'.join(filtered)


def extract_relevant_sections(text: str) -> str:
    """
    ดึงเฉพาะส่วนที่สำคัญสำหรับการประมูลก่อสร้าง
    เพื่อลด token ที่ส่งให้ AI
    """
    lines = text.split('\n')
    scored_lines = []

    for i, line in enumerate(lines):
        score = 0
        line_lower = line.lower()

        # ให้คะแนนบรรทัดที่มี keyword สำคัญ
        for kw in SECTION_KEYWORDS:
            if kw.lower() in line_lower:
                score += 2

        # ให้คะแนน borrow context (บรรทัดใกล้เคียง keyword)
        scored_lines.append((i, score, line))

    # รวม sections ที่มี score สูง + context รอบข้าง
    selected_indices = set()
    for i, score, line in scored_lines:
        if score > 0:
            # เพิ่ม context ±3 บรรทัด
            for j in range(max(0, i-3), min(len(lines), i+4)):
                selected_indices.add(j)

    if not selected_indices:
        # ถ้าหาไม่ได้ ส่งทั้งหมด (truncated)
        return text[:MAX_CHARS]

    result_lines = [lines[i] for i in sorted(selected_indices)]
    result = '\n'.join(result_lines)

    return result[:MAX_CHARS]


# ================================================================
# MAIN FUNCTION
# ================================================================

def preprocess_file(filepath: str | Path) -> dict:
    """
    อ่านไฟล์ TOR แล้วคืน dict พร้อมส่งให้ AI

    Returns:
        {
            "filename": str,
            "file_type": "pdf" | "excel" | "unknown",
            "raw_char_count": int,
            "cleaned_text": str,       ← ส่งให้ AI
            "extraction_ok": bool,
        }
    """
    filepath = Path(filepath)
    result = {
        "filename": filepath.name,
        "file_type": "unknown",
        "raw_char_count": 0,
        "cleaned_text": "",
        "extraction_ok": False,
    }

    if not filepath.exists():
        result["cleaned_text"] = f"[ERROR: ไฟล์ไม่พบ: {filepath}]"
        return result

    suffix = filepath.suffix.lower()

    if suffix == ".pdf":
        result["file_type"] = "pdf"
        raw = read_pdf(filepath)
    elif suffix in (".xlsx", ".xls"):
        result["file_type"] = "excel"
        raw = read_excel(filepath)
    else:
        result["cleaned_text"] = f"[SKIP: ไม่รองรับ format {suffix}]"
        return result

    if raw.startswith("[ERROR"):
        result["cleaned_text"] = raw
        return result

    result["raw_char_count"] = len(raw)
    cleaned = clean_text(raw)
    result["cleaned_text"] = extract_relevant_sections(cleaned)
    result["extraction_ok"] = True

    return result


def preprocess_batch(job_list: list[dict]) -> list[dict]:
    """
    รับ list ของ jobs ที่มี tor_file path แล้วเพิ่ม preprocessed_text เข้าไป
    """
    results = []
    for job in job_list:
        tor_file = job.get("tor_file", "")
        if tor_file and Path(tor_file).exists():
            pp = preprocess_file(tor_file)
            job["preprocessed_text"] = pp["cleaned_text"]
            job["extraction_ok"] = pp["extraction_ok"]
        else:
            job["preprocessed_text"] = f"[ไม่มีไฟล์ TOR — ใช้ข้อมูลจาก sheet: {job.get('title','')}]"
            job["extraction_ok"] = False
        results.append(job)
    return results


# ================================================================
# STANDALONE USAGE
# ================================================================

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: python Sebastian_Preprocessor.py <file.pdf|file.xlsx>")
        sys.exit(1)

    result = preprocess_file(sys.argv[1])
    print(f"\n{'='*60}")
    print(f"ไฟล์:     {result['filename']}")
    print(f"ประเภท:    {result['file_type']}")
    print(f"ขนาดดิบ:  {result['raw_char_count']:,} chars")
    print(f"หลัง trim: {len(result['cleaned_text']):,} chars")
    print(f"{'='*60}")
    print(result['cleaned_text'][:2000])
    if len(result['cleaned_text']) > 2000:
        print(f"\n... (ตัดแสดงที่ 2000 chars จาก {len(result['cleaned_text'])} ทั้งหมด)")
