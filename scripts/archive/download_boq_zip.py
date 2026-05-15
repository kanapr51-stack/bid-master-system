"""
download_boq_zip.py — ดาวน์โหลดและตรวจสอบ BOQ zip file
เพื่อดูว่ามี ปร.4/5 Excel หรือไม่

วิธีใช้: python scripts/download_boq_zip.py
"""
import sys, json, time, base64, zipfile, io
from pathlib import Path
from datetime import datetime
sys.stdout.reconfigure(encoding="utf-8")
from playwright.sync_api import sync_playwright

PROCESS5_BASE = "https://process5.gprocurement.go.th"
SEARCH_URL    = f"{PROCESS5_BASE}/egp-agpc01-web/announcement"
DEBUG_DIR     = Path(__file__).parent.parent / "downloads" / "debug" / "boq_zip"
DEBUG_DIR.mkdir(parents=True, exist_ok=True)

APIKEY = "Liaqv30xLpFGOlJPW1N0hPKJkbO7vWUS"

# บ้านแพง job — has 1.7MB zip
TARGET = {
    "projectId": "69049122041",
    "zipFileId": "ce77ebe946a048969a23324d649fd3be",
    "label": "บ้านแพง ก่อสร้างถนน",
}


def log(msg):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


with sync_playwright() as p:
    log("เชื่อมต่อ Chrome...")
    browser = p.chromium.connect_over_cdp("http://127.0.0.1:9222")
    log("เชื่อมต่อสำเร็จ")

    page = browser.contexts[0].new_page()
    page.goto(SEARCH_URL, wait_until="load", timeout=45000)
    time.sleep(3)
    log("รอ search button...")
    deadline = time.time() + 40
    while time.time() < deadline:
        btn = page.query_selector("button:has-text('ค้นหา')")
        if btn and btn.is_enabled():
            log("  ✓ enabled")
            break
        time.sleep(0.8)

    pid = TARGET["projectId"]
    file_id = TARGET["zipFileId"]

    # ===== 1. ดาวน์โหลด BOQ zip เป็น base64 =====
    log(f"\n=== Download BOQ zip for {pid} ===")
    dl_url = f"{PROCESS5_BASE}/egp-upload-service/v1/downloadFileTest?fileId={file_id}"

    js_download = f"""async () => {{
        try {{
            const r = await fetch({json.dumps(dl_url)}, {{
                credentials: 'include',
            }});
            const buf = await r.arrayBuffer();
            const bytes = new Uint8Array(buf);
            let binary = '';
            for (let i = 0; i < bytes.byteLength; i++) {{
                binary += String.fromCharCode(bytes[i]);
            }}
            return {{
                status: r.status,
                size: buf.byteLength,
                content_type: r.headers.get('content-type') || '',
                content_disposition: r.headers.get('content-disposition') || '',
                b64: btoa(binary)
            }};
        }} catch(e) {{ return {{error: e.toString()}}; }}
    }}"""

    log("  Downloading (this may take a moment)...")
    res = page.evaluate(js_download)
    log(f"  status: {res.get('status')}, size: {res.get('size')}, type: {res.get('content_type')}")
    log(f"  content-disposition: {res.get('content_disposition')}")

    if res.get("b64"):
        zip_bytes = base64.b64decode(res["b64"])
        zip_path = DEBUG_DIR / f"pricebuild_{pid}.zip"
        zip_path.write_bytes(zip_bytes)
        log(f"  บันทึก: {zip_path} ({len(zip_bytes):,} bytes)")

        # ===== 2. ตรวจสอบเนื้อหา zip =====
        log(f"\n=== ตรวจสอบ zip contents ===")
        try:
            with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
                names = zf.namelist()
                log(f"  Files in zip: {len(names)}")
                for name in names:
                    info = zf.getinfo(name)
                    log(f"    {name} ({info.file_size:,} bytes)")

                # Extract and examine key files
                for name in names:
                    ext = Path(name).suffix.lower()
                    if ext in (".xlsx", ".xls", ".xlsm", ".pdf", ".csv"):
                        log(f"\n  ★ ไฟล์สำคัญ: {name}")
                        content = zf.read(name)
                        extract_path = DEBUG_DIR / name.replace("/", "_")
                        extract_path.write_bytes(content)
                        log(f"    บันทึก: {extract_path} ({len(content):,} bytes)")

                        if ext in (".xlsx", ".xls", ".xlsm"):
                            # Try to read Excel
                            try:
                                import openpyxl
                                wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
                                log(f"    Sheets: {wb.sheetnames}")
                                for sheet_name in wb.sheetnames[:3]:
                                    ws = wb[sheet_name]
                                    log(f"    Sheet '{sheet_name}': {ws.max_row} rows x {ws.max_column} cols")
                                    # Print first few rows
                                    for row_idx, row in enumerate(ws.iter_rows(max_row=10, values_only=True)):
                                        non_none = [str(c)[:30] for c in row if c is not None]
                                        if non_none:
                                            log(f"      row {row_idx+1}: {' | '.join(non_none[:8])}")
                            except Exception as e:
                                log(f"    Excel read error: {e}")
                        elif ext == ".pdf":
                            log(f"    (PDF file — will need parser)")
        except Exception as e:
            log(f"  zip error: {e}")
            # Maybe it's not a zip — check what it is
            log(f"  First 20 bytes: {zip_bytes[:20]!r}")
    else:
        log(f"  No b64 data: {res}")

    # ===== 3. ลอง ด่านช้าง job =====
    log(f"\n=== Download ด่านช้าง job ===")
    file_id2 = "25ad05370d3e41bd8fd92b2aa6c2c7fe"
    dl_url2 = f"{PROCESS5_BASE}/egp-upload-service/v1/downloadFileTest?fileId={file_id2}"
    js2 = f"""async () => {{
        const r = await fetch({json.dumps(dl_url2)}, {{credentials: 'include'}});
        const buf = await r.arrayBuffer();
        const bytes = new Uint8Array(buf);
        let binary = '';
        for (let i = 0; i < bytes.byteLength; i++) {{ binary += String.fromCharCode(bytes[i]); }}
        return {{status: r.status, size: buf.byteLength, ct: r.headers.get('content-type')||'', cd: r.headers.get('content-disposition')||'', b64: btoa(binary)}};
    }}"""
    res2 = page.evaluate(js2)
    log(f"  status={res2.get('status')}, size={res2.get('size')}, cd={res2.get('cd')}")
    if res2.get("b64"):
        zip2 = base64.b64decode(res2["b64"])
        zip_path2 = DEBUG_DIR / "pricebuild_69049267400.zip"
        zip_path2.write_bytes(zip2)
        log(f"  บันทึก: {zip_path2} ({len(zip2):,} bytes)")
        try:
            with zipfile.ZipFile(io.BytesIO(zip2)) as zf2:
                names2 = zf2.namelist()
                log(f"  Files in zip: {len(names2)}")
                for n in names2:
                    info2 = zf2.getinfo(n)
                    log(f"    {n} ({info2.file_size:,} bytes)")
        except Exception as e:
            log(f"  zip error: {e}")

    page.close()

log(f"\nดู files ใน: {DEBUG_DIR}")
